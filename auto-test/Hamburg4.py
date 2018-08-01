from bs4 import BeautifulSoup
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def initialize_wb():
	wb = load_workbook("HamburgLists.xlsx")		##Creates and selects the new workbook file
	ws1 = wb.active
	ws1.title = "Hamburg Passenger Lists"
	return wb

def initialize_driver():
	login_url = "https://www.ancestry.com"		##initializes the web driver
	driver = webdriver.Safari()
	driver.get(login_url) ##goes to login page
	return driver

def login(driver):
	WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, 'ancLoginBtn link calloutTrigger iconAfter iconArrowSmallDownAfter')))
	driver.find_element_by_class_name('ancLoginBtn link calloutTrigger iconAfter iconArrowSmallDownAfter').click()
	WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, 'smpSigninBtn')))
	username = driver.find_element_by_id("smpUserName")
	password = driver.find_element_by_id("smpPassword")		##logins into Ancestry.com with appropriate wait times
	username.send_keys("costas.arkolakis@yale.edu")
	password.send_keys("migrationclp!")
	driver.find_element_by_id("smpSigninBtn").click()

def get_to_page(driver, url):
	WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "marketingOverlay")))	##gets to the search results for immigrants to the U.S.
	driver.find_element_by_class_name("closeBtn modalClose").click()
	driver.get(srch_url)

def get_data(driver, row_num, wb):
	WebDriverWait(driver, 15).until(EC.presence_of_element_located((By.ID, "recordData")))
	html_source = driver.page_source					##gets the HTML and table to be parsed
	soup = BeautifulSoup(html_source, "lxml")
	table = soup.find(id = "recordData")
	rows = table.find_all("tr")
	i = 1
	ws1 = wb.active
	for row in rows:
		if row:				
			n = 1
			string = row.get_text()					##breaks the table into rows and breaks each cell into tuples using string.partition()
			str = string.partition(":")
			if row_num == 2:
				max = len(str[0])
				ws1.cell(column = i, row = 1).value = str[0]		##sets the top row of categories (record 2 has all the categories)
				max = len(str[2]) if len(str[2]) > len(str[0]) else max
				ws1.column_dimensions[get_column_letter(i)].width = max
			while n != 20:											##matches the first part of the partition with the category then sets the cell accordingly
				if str[0] == ws1.cell(column = n, row = 1).value:
					ws1.cell(column = n, row = row_num).value = str[2]
					break
				n += 1
			i += 1
	WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CLASS_NAME, "icon iconArrowBack")))			##clicks back to search results
	driver.find_element_by_class_name("icon iconArrowBack").click()	
	wb.save(filename = "HamburgLists.xlsx")


##srch_url = "http://search.ancestry.com/cgi-bin/sse.dll?db=HamburgPL_full&gss=sfs28_ms_db&new=1&rank=1&msT=1&_82004290__ftp=United%20States&_82004290=2&_82004290_PInfo=3-%7C0%7C1652393%7C0%7C2%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C&MSAV=1&MSV=0&uidh=6bk"
srch_url = "http://search.ancestry.com/cgi-bin/sse.dll?_phsrc=BiM1&_phstart=successSource&usePUBJs=true&db=HamburgPL_full&gss=angs-d&new=1&rank=1&msT=1&_82004290__ftp=United%20States&_82004290=2&_82004290_PInfo=3-%7C0%7C1652393%7C0%7C2%7C0%7C0%7C0%7C0%7C0%7C0%7C0%7C&MSAV=1&MSV=0&uidh=6bk&gl=&gst=&hc=50&fh=100&fsk=BEHFl5wIgAAELAAAk8I-61-"
row_num = 123
n = 22
driver = initialize_driver()			##sets up the driver, workbook, and webpage
login(driver)
get_to_page(driver, srch_url) 
wb = initialize_wb()


	
while n < 50:						##loops through the records on the first web page
	WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.ID, "sRes-49")))
	buttons = driver.find_elements_by_class_name("srchFoundDB");
	buttons[n].click()
	get_data(driver, row_num, wb)
	n += 1
	row_num += 1
	if n == 50:
		n = 0
		WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.CLASS_NAME, "ancBtn sml green icon iconArrowRight")))
		driver.find_elements_by_class_name("ancBtn sml green icon iconArrowRight").click()




